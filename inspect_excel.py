import openpyxl

file_name = "Xi măng CEM II A-L 42.5N OMANCO 12.09( Bản chuẩn).xlsx"

wb = openpyxl.load_workbook(file_name, data_only=False)

output_lines = []
output_lines.append("Sheets in workbook: " + str(wb.sheetnames))

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    output_lines.append(f"\n--- Sheet: {sheet_name} (Max row: {sheet.max_row}, Max col: {sheet.max_column}) ---")
    
    # Read first 100 rows and all columns
    for r in range(1, min(sheet.max_row + 1, 100)):
        row_vals = []
        for c in range(1, min(sheet.max_column + 1, 30)):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                if isinstance(val, str) and val.startswith("="):
                    row_vals.append(f"Col {c} (R{r}): {val} (Formula)")
                else:
                    row_vals.append(f"Col {c} (R{r}): '{val}'")
        if row_vals:
            output_lines.append(f"Row {r:2d}: {', '.join(row_vals)}")

with open("excel_structure.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(output_lines))

print("Structure written to excel_structure.txt successfully.")
